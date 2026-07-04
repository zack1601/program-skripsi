#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════╗
║          GEOCODING PROFESIONAL - data_noc1.xlsx                  ║
║          Convert Alamat Indonesia → Koordinat Google Maps        ║
║          Nominatim OpenStreetMap | geopy | pandas                ║
╚══════════════════════════════════════════════════════════════════╝

Cara install dependency:
    pip install pandas openpyxl geopy

Cara menjalankan:
    python geocoding_noc1.py

Input  : data_noc1.xlsx  (harus satu folder dengan script ini)
Output : hasil_geocoding_data_noc1.xlsx
Log    : geocoding.log
"""

import os
import re
import time
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

# ─────────────────────────────────────────────
#  KONFIGURASI
# ─────────────────────────────────────────────
INPUT_FILE       = "data_noc1.xlsx"
OUTPUT_FILE      = "hasil_geocoding_data_noc1.xlsx"
LOG_FILE         = "geocoding.log"
COLUMN_ALAMAT    = "ALAMAT"
AUTO_SAVE_EVERY  = 20       # simpan otomatis setiap N baris
MAX_RETRY        = 3        # maksimal retry jika timeout
RETRY_DELAY      = 2        # detik antar retry
RATE_LIMIT_DELAY = 1.1      # detik antar request ke Nominatim (min 1 detik)

# Kata kunci wilayah Jabodetabek — jika tidak ada, append ", INDONESIA"
WILAYAH_KEYWORDS = [
    "JAKARTA", "DEPOK", "BOGOR", "TANGERANG", "BEKASI"
]

# Prioritas hasil geocoding (diurutkan berdasarkan preferensi)
PRIORITY_KEYWORDS = [
    "PONDOK LABU", "CILANDAK", "JAKARTA SELATAN",
    "CINERE", "DEPOK", "JAKARTA"
]


# ─────────────────────────────────────────────
#  SETUP LOGGING
# ─────────────────────────────────────────────
def setup_logging() -> logging.Logger:
    """Konfigurasi logging ke file dan konsol sekaligus."""
    logger = logging.getLogger("geocoder")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # File handler — semua level
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    # Console handler — INFO ke atas
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


logger = setup_logging()


# ─────────────────────────────────────────────
#  CLEANING ALAMAT
# ─────────────────────────────────────────────
# Peta replacement kata — urutan penting (paling spesifik dulu)
REPLACEMENTS = [
    (r"\bJAKSEL\b",     "JAKARTA SELATAN"),
    (r"\bJAKUT\b",      "JAKARTA UTARA"),
    (r"\bJAKBAR\b",     "JAKARTA BARAT"),
    (r"\bJAKTIM\b",     "JAKARTA TIMUR"),
    (r"\bJAKPUS\b",     "JAKARTA PUSAT"),
    (r"\bPD\.\s*LABU\b","PONDOK LABU"),
    (r"\bPDK\.\s*LABU\b","PONDOK LABU"),
    (r"\bKEL\.\s*",     ""),
    (r"\bKEC\.\s*",     ""),
    (r"\bKEL\b",        ""),
    (r"\bKEC\b",        ""),
    (r"\bJL\.\s*",      "JALAN "),
    (r"\bJLN\.\s*",     "JALAN "),
    (r"\bJLN\b",        "JALAN"),
    (r"\bGG\.\s*",      "GANG "),
    (r"\bRT\b",         "RT"),   # normalkan, bisa distrip nanti
    (r"\bRW\b",         "RW"),
    (r"NO\.\s*",        "NO "),
]


def clean_address(raw: str) -> str:
    """
    Bersihkan dan normalisasi alamat sebelum dikirim ke geocoder.

    Langkah:
    1. Uppercase
    2. Hapus karakter non-ASCII / aneh
    3. Terapkan replacement khusus
    4. Hapus double-space & trim
    5. Tambah ', INDONESIA' jika tidak ada kata wilayah Jabodetabek
    """
    if not isinstance(raw, str) or not raw.strip():
        return ""

    addr = raw.upper().strip()

    # Hapus karakter yang bukan huruf, angka, spasi, koma, titik, /
    addr = re.sub(r"[^\w\s,.\-/]", " ", addr)

    # Terapkan replacement satu per satu
    for pattern, replacement in REPLACEMENTS:
        addr = re.sub(pattern, replacement, addr)

    # Hapus double space
    addr = re.sub(r"\s{2,}", " ", addr).strip()

    # Tambahkan konteks wilayah jika diperlukan
    has_region = any(kw in addr for kw in WILAYAH_KEYWORDS)
    if not has_region:
        addr += ", INDONESIA"

    return addr


# ─────────────────────────────────────────────
#  GEOCODER SETUP
# ─────────────────────────────────────────────
def build_geocoder():
    """
    Buat instance Nominatim dengan RateLimiter.
    user_agent wajib diisi agar tidak diblokir Nominatim.
    """
    geolocator = Nominatim(
        user_agent="geocoding_noc1_app/1.0 (data@internal)",
        timeout=10
    )
    # RateLimiter memastikan minimal RATE_LIMIT_DELAY detik antar request
    geocode = RateLimiter(
        geolocator.geocode,
        min_delay_seconds=RATE_LIMIT_DELAY,
        error_wait_seconds=5,
        max_retries=MAX_RETRY,
        return_value_on_exception=None
    )
    return geocode


# ─────────────────────────────────────────────
#  GEOCODING SINGLE ALAMAT (dengan retry manual)
# ─────────────────────────────────────────────
def geocode_address(geocode_fn, raw_address: str) -> dict:
    """
    Geocode satu alamat. Kembalikan dict dengan semua kolom output.
    Prioritaskan hasil yang mengandung keyword lokasi tertentu.
    """
    empty = {
        "LATITUDE": None,
        "LONGITUDE": None,
        "GOOGLE_MAPS": "",
        "FORMATTED_ADDRESS": "",
        "STATUS": "NOT FOUND",
        "CLEAN_ALAMAT": ""
    }

    if not raw_address or str(raw_address).strip() == "":
        empty["STATUS"] = "EMPTY"
        return empty

    clean = clean_address(str(raw_address))
    empty["CLEAN_ALAMAT"] = clean

    # Coba beberapa variasi query untuk meningkatkan hit-rate
    queries = _build_query_variants(clean)

    for attempt, query in enumerate(queries, 1):
        for retry in range(1, MAX_RETRY + 1):
            try:
                logger.debug(f"  Query #{attempt} (retry {retry}): {query}")
                location = geocode_fn(
                    query,
                    language="id",
                    addressdetails=True,
                    exactly_one=False,   # ambil beberapa kandidat
                    limit=5
                )

                if location:
                    # Pilih kandidat terbaik berdasarkan prioritas
                    best = _pick_best_result(location)
                    lat  = round(best.latitude,  7)
                    lon  = round(best.longitude, 7)
                    return {
                        "LATITUDE":          lat,
                        "LONGITUDE":         lon,
                        "GOOGLE_MAPS":       f"https://maps.google.com/?q={lat},{lon}",
                        "FORMATTED_ADDRESS": best.address,
                        "STATUS":            "VALID",
                        "CLEAN_ALAMAT":      clean
                    }
                break  # tidak error, tapi tidak ketemu → coba query berikutnya

            except (GeocoderTimedOut, GeocoderServiceError) as exc:
                logger.warning(f"  Timeout/Service error (retry {retry}/{MAX_RETRY}): {exc}")
                if retry < MAX_RETRY:
                    time.sleep(RETRY_DELAY)
                else:
                    logger.error(f"  Gagal setelah {MAX_RETRY} retry: {query}")

            except Exception as exc:
                logger.error(f"  Unexpected error: {exc}")
                break

    empty["CLEAN_ALAMAT"] = clean
    return empty


def _build_query_variants(clean: str) -> list[str]:
    """
    Buat beberapa variasi query dari alamat bersih.
    Query pertama paling lengkap, berikutnya lebih sederhana.
    """
    variants = [clean]

    # Tambah "Jakarta" jika belum ada namun ada konteks Selatan/dll
    if "JAKARTA" not in clean and any(
        x in clean for x in ["SELATAN", "UTARA", "BARAT", "TIMUR", "PUSAT"]
    ):
        variants.append(clean.replace("SELATAN", "JAKARTA SELATAN")
                             .replace("UTARA",   "JAKARTA UTARA")
                             .replace("BARAT",   "JAKARTA BARAT")
                             .replace("TIMUR",   "JAKARTA TIMUR")
                             .replace("PUSAT",   "JAKARTA PUSAT"))

    # Versi tanpa RT/RW (sering mengacaukan geocoder)
    no_rtrw = re.sub(r"\bRT\s*\d+[/\-]?\s*(RW\s*\d+)?\b", "", clean)
    no_rtrw = re.sub(r"\bRW\s*\d+\b", "", no_rtrw)
    no_rtrw = re.sub(r"\s{2,}", " ", no_rtrw).strip()
    if no_rtrw != clean:
        variants.append(no_rtrw)

    # Versi tanpa nomor jalan
    no_no = re.sub(r"\bNO\s*\d+[A-Z]?\b", "", no_rtrw)
    no_no = re.sub(r"\s{2,}", " ", no_no).strip()
    if no_no != no_rtrw:
        variants.append(no_no)

    return variants


def _pick_best_result(candidates):
    """
    Pilih kandidat geocoding terbaik berdasarkan PRIORITY_KEYWORDS.
    Jika tidak ada yang cocok, kembalikan kandidat pertama.
    """
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    for keyword in PRIORITY_KEYWORDS:
        for candidate in candidates:
            if keyword in candidate.address.upper():
                return candidate

    return candidates[0]


# ─────────────────────────────────────────────
#  PROGRESS DISPLAY
# ─────────────────────────────────────────────
def print_progress(idx: int, total: int, address: str, success: int, failed: int):
    """Tampilkan progress ke konsol dengan format rapi."""
    pct   = (idx / total) * 100
    bar_w = 30
    filled = int(bar_w * idx / total)
    bar   = "█" * filled + "░" * (bar_w - filled)

    addr_display = address[:60] + "…" if len(address) > 60 else address
    print(
        f"\r[{idx:>5}/{total}] |{bar}| {pct:5.1f}%  "
        f"✓{success}  ✗{failed}  "
        f"→ {addr_display:<63}",
        end="", flush=True
    )


# ─────────────────────────────────────────────
#  AUTO SAVE
# ─────────────────────────────────────────────
def auto_save(df: pd.DataFrame, output_path: str, idx: int):
    """Simpan DataFrame ke file Excel sementara setiap AUTO_SAVE_EVERY baris."""
    try:
        df.to_excel(output_path, index=False, engine="openpyxl")
        logger.debug(f"Auto-save pada baris {idx} → {output_path}")
    except Exception as exc:
        logger.warning(f"Auto-save gagal pada baris {idx}: {exc}")


# ─────────────────────────────────────────────
#  FORMAT OUTPUT EXCEL
# ─────────────────────────────────────────────
def format_output_excel(df: pd.DataFrame, output_path: str):
    """
    Simpan DataFrame ke Excel dengan formatting profesional
    menggunakan openpyxl styling.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    # Simpan dulu dengan pandas
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active

    # ── Warna header ──
    HEADER_FILL  = PatternFill("solid", start_color="1F4E79")   # biru tua
    VALID_FILL   = PatternFill("solid", start_color="C6EFCE")   # hijau muda
    NOTFOUND_FILL= PatternFill("solid", start_color="FFCCCC")   # merah muda
    ALT_FILL     = PatternFill("solid", start_color="EBF3FB")   # biru sangat muda

    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    CELL_FONT    = Font(name="Arial", size=9)
    LINK_FONT    = Font(name="Arial", size=9, color="0563C1", underline="single")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Format header row ──
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    ws.row_dimensions[1].height = 28

    # ── Tentukan indeks kolom STATUS & GOOGLE_MAPS ──
    headers = [cell.value for cell in ws[1]]
    status_col = headers.index("STATUS") + 1 if "STATUS" in headers else None
    maps_col   = headers.index("GOOGLE_MAPS") + 1 if "GOOGLE_MAPS" in headers else None
    lat_col    = headers.index("LATITUDE") + 1 if "LATITUDE" in headers else None
    lon_col    = headers.index("LONGITUDE") + 1 if "LONGITUDE" in headers else None

    # ── Format data rows ──
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_alt = row_idx % 2 == 0
        status_val = ws.cell(row=row_idx, column=status_col).value if status_col else ""

        for cell in row:
            cell.font   = CELL_FONT
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            # Warna baris berdasarkan status
            if status_col and cell.column == status_col:
                if status_val == "VALID":
                    cell.fill = VALID_FILL
                    cell.font = Font(name="Arial", size=9, bold=True, color="276221")
                elif status_val == "NOT FOUND":
                    cell.fill = NOTFOUND_FILL
                    cell.font = Font(name="Arial", size=9, bold=True, color="9C0006")
            elif is_alt:
                cell.fill = ALT_FILL

            # Format angka lat/lon
            if lat_col and cell.column == lat_col:
                cell.number_format = "0.0000000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if lon_col and cell.column == lon_col:
                cell.number_format = "0.0000000"
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Google Maps link — hyperlink
            if maps_col and cell.column == maps_col and cell.value:
                url = str(cell.value)
                cell.hyperlink = url
                cell.font = LINK_FONT

    # ── Lebar kolom otomatis ──
    col_widths = {
        "ALAMAT": 45, "CLEAN_ALAMAT": 45,
        "LATITUDE": 14, "LONGITUDE": 14,
        "GOOGLE_MAPS": 40, "FORMATTED_ADDRESS": 50,
        "STATUS": 12
    }
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width  = col_widths.get(header, 20)
        ws.column_dimensions[letter].width = width

    # ── Freeze pane di baris ke-2 ──
    ws.freeze_panes = "A2"

    # ── Auto filter ──
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info(f"Excel tersimpan dengan formatting: {output_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    start_time = datetime.now()

    print("=" * 70)
    print("  GEOCODING PROFESIONAL — data_noc1.xlsx")
    print(f"  Mulai: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    logger.info("=" * 60)
    logger.info("Geocoding dimulai")

    # ── 1. Cek file input ──
    script_dir  = Path(__file__).parent
    input_path  = script_dir / INPUT_FILE
    output_path = script_dir / OUTPUT_FILE

    if not input_path.exists():
        logger.error(f"File tidak ditemukan: {input_path}")
        print(f"\n❌  ERROR: File '{INPUT_FILE}' tidak ditemukan.")
        print(f"    Pastikan file berada di folder yang sama dengan script ini:")
        print(f"    {script_dir}")
        return

    # ── 2. Baca Excel ──
    try:
        logger.info(f"Membaca file: {input_path}")
        df = pd.read_excel(input_path, dtype=str, engine="openpyxl")
        df.columns = df.columns.str.strip().str.upper()
        logger.info(f"Jumlah baris: {len(df)} | Kolom: {list(df.columns)}")
    except Exception as exc:
        logger.error(f"Gagal membaca file Excel: {exc}")
        print(f"\n❌  ERROR membaca file: {exc}")
        return

    # ── 3. Validasi kolom ALAMAT ──
    if COLUMN_ALAMAT not in df.columns:
        available = ", ".join(df.columns)
        logger.error(f"Kolom '{COLUMN_ALAMAT}' tidak ditemukan. Kolom tersedia: {available}")
        print(f"\n❌  ERROR: Kolom '{COLUMN_ALAMAT}' tidak ada di file.")
        print(f"    Kolom yang tersedia: {available}")
        return

    total = len(df)
    print(f"\n📄  File     : {INPUT_FILE}")
    print(f"📊  Total    : {total:,} baris")
    print(f"💾  Output   : {OUTPUT_FILE}")
    print(f"📝  Log      : {LOG_FILE}")
    print("-" * 70)

    # ── 4. Inisialisasi kolom output ──
    for col in ["LATITUDE", "LONGITUDE", "GOOGLE_MAPS", "FORMATTED_ADDRESS", "STATUS", "CLEAN_ALAMAT"]:
        if col not in df.columns:
            df[col] = None

    # ── 5. Setup geocoder ──
    geocode_fn = build_geocoder()
    logger.info("Geocoder Nominatim siap")

    # ── 6. Loop utama ──
    success_count = 0
    failed_count  = 0

    print("\n🌐  Memulai proses geocoding...\n")

    for idx, row in df.iterrows():
        row_num   = idx + 1
        raw_addr  = str(row.get(COLUMN_ALAMAT, "")).strip()

        result = geocode_address(geocode_fn, raw_addr)

        # Tulis hasil ke DataFrame
        df.at[idx, "LATITUDE"]          = result["LATITUDE"]
        df.at[idx, "LONGITUDE"]         = result["LONGITUDE"]
        df.at[idx, "GOOGLE_MAPS"]       = result["GOOGLE_MAPS"]
        df.at[idx, "FORMATTED_ADDRESS"] = result["FORMATTED_ADDRESS"]
        df.at[idx, "STATUS"]            = result["STATUS"]
        df.at[idx, "CLEAN_ALAMAT"]      = result["CLEAN_ALAMAT"]

        if result["STATUS"] == "VALID":
            success_count += 1
            logger.info(f"[{row_num}/{total}] VALID   | {raw_addr[:60]}")
            logger.debug(f"  → {result['FORMATTED_ADDRESS']}")
            logger.debug(f"  → {result['LATITUDE']}, {result['LONGITUDE']}")
        else:
            failed_count += 1
            logger.warning(f"[{row_num}/{total}] {result['STATUS']:10s} | {raw_addr[:60]}")

        # Tampilkan progress
        print_progress(row_num, total, raw_addr, success_count, failed_count)

        # Auto save setiap AUTO_SAVE_EVERY baris
        if row_num % AUTO_SAVE_EVERY == 0:
            print()  # newline sebelum log auto-save
            auto_save(df, str(output_path), row_num)

    print()  # newline setelah progress bar selesai

    # ── 7. Simpan final dengan formatting ──
    print("\n💾  Menyimpan file output dengan formatting...")
    try:
        format_output_excel(df, str(output_path))
    except Exception as exc:
        logger.error(f"Gagal format Excel: {exc}. Coba simpan tanpa format...")
        df.to_excel(str(output_path), index=False, engine="openpyxl")
        logger.info(f"Tersimpan (tanpa format): {output_path}")

    # ── 8. Summary ──
    end_time  = datetime.now()
    duration  = end_time - start_time
    total_sec = int(duration.total_seconds())
    hours, rem = divmod(total_sec, 3600)
    minutes, seconds = divmod(rem, 60)
    dur_str   = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    pct_valid = (success_count / total * 100) if total > 0 else 0

    summary = f"""
{'='*70}
  ✅  GEOCODING SELESAI — RINGKASAN
{'='*70}
  📊  Total data    : {total:>8,}
  ✅  Berhasil      : {success_count:>8,}  ({pct_valid:.1f}%)
  ❌  Gagal         : {failed_count:>8,}  ({100 - pct_valid:.1f}%)
  ⏱️  Durasi        : {dur_str}
  💾  Output        : {output_path}
  📝  Log file      : {script_dir / LOG_FILE}
{'='*70}
"""
    print(summary)
    logger.info("=" * 60)
    logger.info(f"SELESAI | Total: {total} | Valid: {success_count} | Gagal: {failed_count} | Durasi: {dur_str}")
    logger.info(f"Output disimpan di: {output_path}")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    main()